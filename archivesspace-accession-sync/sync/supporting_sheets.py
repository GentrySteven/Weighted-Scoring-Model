"""
Supporting Sheet Writers

Computes and writes the four types of supporting sheets that appear
in the main workbook alongside the "Accession Data and Scores" sheet:

- Backlog Change Over Time: monthly trend data
- Processing Projects Over Time: monthly completion data
- Backlog At a Glance: status-group snapshot dashboard
- Processing Queue sheets: one per configured queue

On every sync the program reads the current main-sheet rows, recomputes
all four supporting sheets from scratch, and writes them back. This
keeps the supporting sheets in sync with the data without relying on
spreadsheet-side formulas or pivot-table refresh.

These functions were previously in sync/cli.py but were extracted
here to keep that module focused on entry points and argument parsing.
"""

from datetime import datetime
from pathlib import Path

from sync.config_manager import ConfigManager
from sync.logging_manager import LoggingManager
from sync.sync_engine import SyncEngine
from sync.visualizations import (
    CompletionByStatusTableBuilder,
    MonthlyChangeTableBuilder,
    SubjectDescriptorCountsBuilder,
    VIZ_COMPLETION_TABLE,
    VIZ_MONTHLY_TABLE,
    VIZ_SHEET,
    VIZ_SUBJECTS_TABLE,
    backlog_rows_for_subjects,
    cells_to_values_array as viz_cells_to_values_array,
    chart_specs,
)


def update_supporting_sheets(
    spreadsheet, sync_engine: SyncEngine, config: ConfigManager,
    logger: LoggingManager, output_format: str,
) -> None:
    """
    Update all supporting sheets with computed values.

    Updates four types of sheets:
    - Backlog Change Over Time (monthly trend, existing)
    - Processing Projects Over Time (monthly completion, existing)
    - Backlog At a Glance (status snapshot dashboard, new)
    - Processing Queue sheets (donor-grouped project lists, new — one per
      configured queue in data.yml)

    Each sheet is recomputed from the current main sheet rows on every sync.
    """
    try:
        # Read current main sheet data
        headers, rows = spreadsheet.read_data()

        # Determine the tool's start date from the first cache or log
        cache_dir = config.get("cache", "directory", default="")
        start_date = datetime.now()
        if cache_dir:
            cache_path = Path(cache_dir) / "accession_cache.json"
            if cache_path.exists():
                try:
                    start_date = datetime.fromtimestamp(cache_path.stat().st_mtime)
                except (OSError, IOError):
                    pass

        # Compute existing supporting sheet data
        backlog_data = sync_engine.compute_backlog_change_over_time(rows, start_date)
        project_data = sync_engine.compute_processing_projects_over_time(rows, start_date)

        # Compute new supporting sheet data
        from sync.processing_queue import (
            BacklogAtAGlanceBuilder, ProcessingQueueBuilder, queue_sheet_name,
        )

        # Backlog At a Glance — uses status groups from data.yml.
        # The project count function counts unique values of the configured
        # grouping field (default: Donor Name) for the general backlog rows.
        bag_config = config.get_data("backlog_at_a_glance", default={})
        status_groups = bag_config.get("status_groups", [])

        # Determine grouping field for project counting (use first queue's
        # grouping field, or fall back to "Donor Name")
        pq_config = config.get_data("processing_queue", default={})
        queues = pq_config.get("queues", [])
        grouping_field = (
            queues[0].get("grouping_field", "Donor Name") if queues else "Donor Name"
        )

        def project_count_func(filtered_rows: list) -> int:
            """Count unique grouping field values in the filtered rows."""
            return len({
                str(r.get(grouping_field, "")).strip()
                for r in filtered_rows
                if r.get(grouping_field)
            })

        bag_builder = BacklogAtAGlanceBuilder(rows, status_groups, project_count_func)
        bag_cells = bag_builder.build_cells() if status_groups else []

        # Processing Queue sheets — one per configured queue
        queue_data: list[tuple[str, list]] = []
        for queue_cfg in queues:
            builder = ProcessingQueueBuilder(rows, queue_cfg)
            cells = builder.build_cells()
            sheet_name = queue_sheet_name(queue_cfg.get("name", "Queue"))
            queue_data.append((sheet_name, cells))

        # Visualization data tables — three hidden sheets that feed the
        # chart objects on the Visualizations sheet. Computed on every
        # sync so charts reflect current data.
        monthly_cells = MonthlyChangeTableBuilder(rows).build_cells()
        completion_cells = CompletionByStatusTableBuilder(
            rows, status_groups
        ).build_cells() if status_groups else []
        # For the Top-N subjects chart we count across rows matching the
        # first queue's status filter (typically "Backlog - General").
        # This matches the original sheet's behavior of counting only
        # backlog accessions, not all accessions.
        backlog_statuses = queues[0].get("status_values", []) if queues else []
        subject_rows = backlog_rows_for_subjects(rows, backlog_statuses)
        subject_cells = SubjectDescriptorCountsBuilder(subject_rows).build_cells()

        viz_data = {
            "monthly": monthly_cells,
            "completion": completion_cells,
            "subjects": subject_cells,
            "num_status_groups": len(status_groups),
        }

        if output_format == "excel":
            _write_supporting_sheets_excel(
                spreadsheet, backlog_data, project_data, logger,
                bag_cells=bag_cells, queue_data=queue_data, viz_data=viz_data,
            )
        elif output_format == "google_sheets":
            _write_supporting_sheets_google(
                spreadsheet, backlog_data, project_data, logger,
                bag_cells=bag_cells, queue_data=queue_data, viz_data=viz_data,
            )

        logger.technical(
            f"Supporting sheets updated: {len(backlog_data)} backlog months, "
            f"{len(queues)} processing queue(s), "
            f"{len(status_groups)} status groups, "
            f"visualizations ({len(monthly_cells)} monthly rows, "
            f"{len(completion_cells)} completion rows, "
            f"{len(subject_cells)} subject rows)."
        )

    except Exception as e:
        logger.warning(f"Failed to update supporting sheets: {e}")


def _write_supporting_sheets_excel(
    spreadsheet, backlog_data, project_data, logger,
    bag_cells: list = None, queue_data: list = None,
    viz_data: dict = None,
):
    """
    Write computed supporting sheet data to Excel.

    Args:
        bag_cells: Cell tuples for the Backlog At a Glance sheet.
        queue_data: List of (sheet_name, cells) for each processing queue.
    """
    try:
        from sync.processing_queue import BacklogAtAGlanceBuilder

        # If the Excel manager is inside a transaction, reuse its workbook
        # and skip the save (the transaction's __exit__ will save once).
        # Otherwise load and save our own workbook, preserving the prior
        # behavior for non-transactional callers.
        in_transaction = getattr(spreadsheet, "_active_wb", None) is not None
        if in_transaction:
            wb = spreadsheet._active_wb
        else:
            import openpyxl
            wb = openpyxl.load_workbook(str(spreadsheet.get_file_path()))

        # Backlog Change Over Time
        if "Backlog Change Over Time" in wb.sheetnames:
            ws = wb["Backlog Change Over Time"]
            # Clear existing data (keep headers)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

            for idx, month_data in enumerate(backlog_data, 2):
                ws.cell(row=idx, column=1, value=month_data["Month and Year"])
                ws.cell(row=idx, column=2, value=month_data["Physical Backlog (Linear Feet)"])
                ws.cell(row=idx, column=3, value=month_data["Digital Backlog (GB)"])

        # Processing Projects Over Time
        if "Processing Projects Over Time" in wb.sheetnames:
            ws = wb["Processing Projects Over Time"]
            # Clear entire sheet including headers (they're dynamic)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.value = None

            # Collect all unique project types
            all_project_types: list[str] = []
            for month_data in project_data:
                for pt in month_data.get("project_counts", {}):
                    if pt not in all_project_types:
                        all_project_types.append(pt)

            # Write header row
            ws.cell(row=1, column=1, value="Month and Year")
            col = 2
            for pt in all_project_types:
                ws.cell(row=1, column=col, value=f"{pt} (Count)")
                ws.cell(row=1, column=col + 1, value=f"{pt} (Physical LF)")
                ws.cell(row=1, column=col + 2, value=f"{pt} (Digital GB)")
                col += 3

            # Write data rows
            for idx, month_data in enumerate(project_data, 2):
                ws.cell(row=idx, column=1, value=month_data["Month and Year"])
                counts = month_data.get("project_counts", {})
                physical = month_data.get("project_physical", {})
                digital = month_data.get("project_digital", {})

                col = 2
                for pt in all_project_types:
                    ws.cell(row=idx, column=col, value=counts.get(pt, 0))
                    ws.cell(row=idx, column=col + 1, value=physical.get(pt, 0))
                    ws.cell(row=idx, column=col + 2, value=digital.get(pt, 0))
                    col += 3

        # Backlog At a Glance — write computed status summary cells
        if bag_cells:
            sheet_name = BacklogAtAGlanceBuilder.SHEET_NAME
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
            # Clear existing content
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.value = None
            # Write new cells
            for row, col, value in bag_cells:
                ws.cell(row=row, column=col, value=value)

        # Processing Queue sheets — one per configured queue
        if queue_data:
            for sheet_name, cells in queue_data:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(title=sheet_name)
                # Clear existing content
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                    for cell in row:
                        cell.value = None
                # Write new cells
                for row, col, value in cells:
                    ws.cell(row=row, column=col, value=value)

        # Visualization data tables (hidden) and chart objects
        if viz_data:
            _write_visualizations_excel(wb, viz_data, logger)

        # Save only if we own this workbook. In a transaction, the
        # ExcelManager's context manager saves on exit.
        if not in_transaction:
            wb.save(str(spreadsheet.get_file_path()))

    except Exception as e:
        logger.warning(f"Failed to write Excel supporting sheets: {e}")


def _write_visualizations_excel(wb, viz_data: dict, logger) -> None:
    """
    Write the three visualization data tables (hidden) and ensure the
    Visualizations sheet exists with the 8 chart objects anchored.

    Charts are created once on first run and persist across subsequent
    syncs. On each sync we refresh the data tables; the charts' data
    references auto-extend to cover any new rows because we re-create
    the chart objects referencing the current extents.
    """
    from sync.visualizations import (
        VIZ_SHEET, VIZ_MONTHLY_TABLE, VIZ_COMPLETION_TABLE, VIZ_SUBJECTS_TABLE,
        chart_specs,
    )

    # --- Step 1: populate/refresh the three hidden data sheets ---
    for sheet_name, cells in [
        (VIZ_MONTHLY_TABLE, viz_data.get("monthly") or []),
        (VIZ_COMPLETION_TABLE, viz_data.get("completion") or []),
        (VIZ_SUBJECTS_TABLE, viz_data.get("subjects") or []),
    ]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.value = None
        else:
            ws = wb.create_sheet(title=sheet_name)
        # Hide the data sheet — these are inputs to charts, not for direct viewing
        ws.sheet_state = "hidden"
        for row, col, value in cells:
            ws.cell(row=row, column=col, value=value)

    # --- Step 2: create/refresh the Visualizations sheet and chart objects ---
    if VIZ_SHEET in wb.sheetnames:
        viz_ws = wb[VIZ_SHEET]
        # Drop existing charts — we rebuild them each sync so data ranges
        # pick up any new rows. openpyxl stores charts in the worksheet's
        # _charts list; clearing this list removes them from the output.
        viz_ws._charts = []
    else:
        viz_ws = wb.create_sheet(title=VIZ_SHEET)
        viz_ws.cell(row=1, column=1, value="Visualizations")
        viz_ws.cell(row=2, column=1, value=(
            "Charts below are regenerated on every sync from the hidden "
            "Viz - data sheets. Do not edit those sheets directly."
        ))

    # Determine the actual last row in each data sheet so the chart's
    # series ranges cover exactly the data that exists.
    data_extents = {}
    for sheet_name in (VIZ_MONTHLY_TABLE, VIZ_COMPLETION_TABLE, VIZ_SUBJECTS_TABLE):
        if sheet_name in wb.sheetnames:
            ds = wb[sheet_name]
            # max_row includes header; a table with only a header has max_row=1
            data_extents[sheet_name] = ds.max_row

    # --- Step 3: build and add each chart ---
    num_groups = viz_data.get("num_status_groups", 0)
    specs = chart_specs(num_status_groups=num_groups)
    for spec in specs:
        chart = _build_chart_from_spec(wb, spec, data_extents)
        if chart is not None:
            viz_ws.add_chart(chart, spec["anchor"])

    logger.technical(f"Wrote {len(specs)} chart(s) to '{VIZ_SHEET}' sheet.")


def _build_chart_from_spec(wb, spec: dict, data_extents: dict):
    """
    Construct an openpyxl chart object from a spec dict.

    Returns None if the source data sheet has no data rows (header only),
    so the chart is skipped rather than emitted empty.
    """
    from openpyxl.chart import BarChart, PieChart, Reference

    table_name = spec["table"]
    last_row = data_extents.get(table_name, 1)
    if last_row < 2:
        # Header only — no data, skip chart
        return None

    data_ws = wb[table_name]
    cat_info = spec["categories"]
    cat_first = cat_info["first_row"]
    cat_last = cat_info["last_row"] if cat_info["last_row"] is not None else last_row

    # Bound the pie chart's last row to the actual data extent
    cat_last = min(cat_last, last_row)
    if cat_last < cat_first:
        return None

    kind = spec["kind"]
    if kind == "pie":
        chart = PieChart()
    else:
        chart = BarChart()
        chart.type = "col"
        if kind == "column_stacked":
            chart.grouping = "stacked"
            chart.overlap = 100
        else:
            chart.grouping = "clustered"
        if spec.get("x_axis"):
            chart.x_axis.title = spec["x_axis"]
        if spec.get("y_axis"):
            chart.y_axis.title = spec["y_axis"]

    chart.title = spec["title"]
    chart.style = 10  # Provides a readable default palette

    # Categories (x-axis labels for column charts, slice labels for pie)
    categories = Reference(
        data_ws,
        min_col=cat_info["col"],
        min_row=cat_first,
        max_col=cat_info["col"],
        max_row=cat_last,
    )
    chart.set_categories(categories)

    # Series: each series reads one column. We read the header row through
    # the data rows so openpyxl picks up the series name from row 1.
    for series_spec in spec["series"]:
        col = series_spec["col"]
        ref = Reference(
            data_ws,
            min_col=col,
            min_row=series_spec.get("name_header_row", 1),
            max_col=col,
            max_row=cat_last,
        )
        chart.add_data(ref, titles_from_data=True)

    return chart


def _write_supporting_sheets_google(
    spreadsheet, backlog_data, project_data, logger,
    bag_cells: list = None, queue_data: list = None,
    viz_data: dict = None,
):
    """
    Write computed supporting sheet data to Google Sheets.

    Args:
        bag_cells: Cell tuples for the Backlog At a Glance sheet.
        queue_data: List of (sheet_name, cells) for each processing queue.
    """
    import time
    from sync.processing_queue import BacklogAtAGlanceBuilder

    try:
        sheets_service = spreadsheet.get_sheets_service()
        spreadsheet_id = spreadsheet.get_spreadsheet_id()
        throttle = spreadsheet.throttle_sheets

        if not sheets_service or not spreadsheet_id:
            logger.warning("Google Sheets service not available for supporting sheets.")
            return

        # Get existing sheet names
        time.sleep(throttle)
        metadata = sheets_service.spreadsheets().get(
            spreadsheetId=spreadsheet_id
        ).execute()
        existing_sheets = {
            s["properties"]["title"]: s["properties"]["sheetId"]
            for s in metadata.get("sheets", [])
        }

        # Create sheets if they don't exist (existing + new)
        required_sheets = [
            "Backlog Change Over Time",
            "Processing Projects Over Time",
            BacklogAtAGlanceBuilder.SHEET_NAME,
        ]
        # Add processing queue sheets to required list
        if queue_data:
            for sheet_name, _ in queue_data:
                required_sheets.append(sheet_name)
        # Add visualization sheets to required list
        if viz_data:
            from sync.visualizations import (
                VIZ_SHEET, VIZ_MONTHLY_TABLE, VIZ_COMPLETION_TABLE, VIZ_SUBJECTS_TABLE,
            )
            required_sheets.extend([
                VIZ_SHEET, VIZ_MONTHLY_TABLE, VIZ_COMPLETION_TABLE, VIZ_SUBJECTS_TABLE,
            ])

        sheets_to_create = [s for s in required_sheets if s not in existing_sheets]

        if sheets_to_create:
            requests = [
                {
                    "addSheet": {
                        "properties": {"title": name}
                    }
                }
                for name in sheets_to_create
            ]
            time.sleep(throttle)
            sheets_service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={"requests": requests},
            ).execute()
            logger.technical(f"Created supporting sheets: {', '.join(sheets_to_create)}")

        # --- Backlog Change Over Time ---
        if backlog_data:
            backlog_values = [
                ["Month and Year", "Physical Backlog (Linear Feet)", "Digital Backlog (GB)"]
            ]
            for month_data in backlog_data:
                backlog_values.append([
                    month_data["Month and Year"],
                    month_data["Physical Backlog (Linear Feet)"],
                    month_data["Digital Backlog (GB)"],
                ])

            # Clear existing data
            time.sleep(throttle)
            sheets_service.spreadsheets().values().clear(
                spreadsheetId=spreadsheet_id,
                range="'Backlog Change Over Time'",
            ).execute()

            # Write new data (headers + rows)
            time.sleep(throttle)
            sheets_service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range="'Backlog Change Over Time'!A1",
                valueInputOption="RAW",
                body={"values": backlog_values},
            ).execute()

            logger.technical(
                f"Backlog Change Over Time: wrote {len(backlog_values) - 1} months."
            )

        # --- Processing Projects Over Time ---
        if project_data:
            # Collect all unique project types across all months
            all_project_types: list[str] = []
            for month_data in project_data:
                for pt in month_data.get("project_counts", {}):
                    if pt not in all_project_types:
                        all_project_types.append(pt)

            # Build header row
            project_headers = ["Month and Year"]
            for pt in all_project_types:
                project_headers.append(f"{pt} (Count)")
                project_headers.append(f"{pt} (Physical LF)")
                project_headers.append(f"{pt} (Digital GB)")

            project_values = [project_headers]

            for month_data in project_data:
                row = [month_data["Month and Year"]]
                counts = month_data.get("project_counts", {})
                physical = month_data.get("project_physical", {})
                digital = month_data.get("project_digital", {})

                for pt in all_project_types:
                    row.append(counts.get(pt, 0))
                    row.append(physical.get(pt, 0))
                    row.append(digital.get(pt, 0))

                project_values.append(row)

            # Clear existing data
            time.sleep(throttle)
            sheets_service.spreadsheets().values().clear(
                spreadsheetId=spreadsheet_id,
                range="'Processing Projects Over Time'",
            ).execute()

            # Write new data
            time.sleep(throttle)
            sheets_service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range="'Processing Projects Over Time'!A1",
                valueInputOption="RAW",
                body={"values": project_values},
            ).execute()

            logger.technical(
                f"Processing Projects Over Time: wrote {len(project_values) - 1} months, "
                f"{len(all_project_types)} project type(s)."
            )

        # --- Backlog At a Glance ---
        if bag_cells:
            sheet_name = BacklogAtAGlanceBuilder.SHEET_NAME
            bag_values = _cells_to_values_array(bag_cells)
            time.sleep(throttle)
            sheets_service.spreadsheets().values().clear(
                spreadsheetId=spreadsheet_id,
                range=f"'{sheet_name}'",
            ).execute()
            time.sleep(throttle)
            sheets_service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range=f"'{sheet_name}'!A1",
                valueInputOption="RAW",
                body={"values": bag_values},
            ).execute()
            logger.technical(f"Backlog At a Glance: wrote {len(bag_values)} rows.")

        # --- Processing Queue sheets ---
        if queue_data:
            for sheet_name, cells in queue_data:
                queue_values = _cells_to_values_array(cells)
                time.sleep(throttle)
                sheets_service.spreadsheets().values().clear(
                    spreadsheetId=spreadsheet_id,
                    range=f"'{sheet_name}'",
                ).execute()
                time.sleep(throttle)
                sheets_service.spreadsheets().values().update(
                    spreadsheetId=spreadsheet_id,
                    range=f"'{sheet_name}'!A1",
                    valueInputOption="RAW",
                    body={"values": queue_values},
                ).execute()
                logger.technical(
                    f"{sheet_name}: wrote {len(queue_values)} rows."
                )

        # --- Visualization data tables and chart objects ---
        if viz_data:
            _write_visualizations_google(
                sheets_service, spreadsheet_id, existing_sheets, viz_data,
                logger, throttle,
            )

    except Exception as e:
        logger.warning(f"Failed to write Google Sheets supporting sheets: {e}")


def _write_visualizations_google(
    sheets_service, spreadsheet_id: str, existing_sheets: dict,
    viz_data: dict, logger, throttle: float,
) -> None:
    """
    Write visualization data tables and chart objects to Google Sheets.

    Unlike Excel, Google Sheets chart objects are tied to a specific
    chartId that persists across updates. On first run we create the
    charts via addChart requests; on subsequent runs we refresh the
    underlying data tables — the charts auto-update because they
    reference ranges that include any new rows (pie chart top-10 stays
    at 10, time-series charts extend). We re-fetch the current chart
    list on each run and only add charts that don't already exist.
    """
    import time
    from sync.visualizations import (
        VIZ_SHEET, VIZ_MONTHLY_TABLE, VIZ_COMPLETION_TABLE, VIZ_SUBJECTS_TABLE,
        chart_specs, cells_to_values_array,
    )

    # --- Step 1: refresh hidden data tables ---
    for sheet_name, cells in [
        (VIZ_MONTHLY_TABLE, viz_data.get("monthly") or []),
        (VIZ_COMPLETION_TABLE, viz_data.get("completion") or []),
        (VIZ_SUBJECTS_TABLE, viz_data.get("subjects") or []),
    ]:
        values = cells_to_values_array(cells)
        if not values:
            continue
        time.sleep(throttle)
        sheets_service.spreadsheets().values().clear(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_name}'",
        ).execute()
        time.sleep(throttle)
        sheets_service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_name}'!A1",
            valueInputOption="RAW",
            body={"values": values},
        ).execute()

    # Hide the three data sheets and ensure the Visualizations sheet is visible.
    # Also fetch current chart IDs to avoid creating duplicates.
    time.sleep(throttle)
    meta = sheets_service.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        includeGridData=False,
    ).execute()

    sheet_by_name = {s["properties"]["title"]: s for s in meta.get("sheets", [])}
    hide_requests: list[dict] = []
    for hidden_name in (VIZ_MONTHLY_TABLE, VIZ_COMPLETION_TABLE, VIZ_SUBJECTS_TABLE):
        s = sheet_by_name.get(hidden_name)
        if s and not s["properties"].get("hidden", False):
            hide_requests.append({
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": s["properties"]["sheetId"],
                        "hidden": True,
                    },
                    "fields": "hidden",
                }
            })
    if hide_requests:
        time.sleep(throttle)
        sheets_service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": hide_requests},
        ).execute()

    # Check if the Visualizations sheet already has charts.
    # If so, we leave them alone — Google Sheets charts auto-refresh
    # when their source ranges' data changes. On first run, the sheet
    # has no charts and we create them.
    viz_sheet = sheet_by_name.get(VIZ_SHEET)
    if not viz_sheet:
        logger.warning(f"Visualizations sheet '{VIZ_SHEET}' missing.")
        return

    existing_charts = viz_sheet.get("charts", [])
    if existing_charts:
        logger.technical(
            f"Visualizations sheet has {len(existing_charts)} existing chart(s); "
            f"skipping chart creation (data ranges refreshed automatically)."
        )
        return

    # --- Step 2: build addChart requests from spec list ---
    num_groups = viz_data.get("num_status_groups", 0)
    specs = chart_specs(num_status_groups=num_groups)

    # Determine actual data extents so charts reference correct ranges
    extents: dict[str, int] = {}
    for t_name in (VIZ_MONTHLY_TABLE, VIZ_COMPLETION_TABLE, VIZ_SUBJECTS_TABLE):
        s = sheet_by_name.get(t_name)
        extents[t_name] = _sheet_row_count(sheets_service, spreadsheet_id, t_name, throttle)

    viz_sheet_id = viz_sheet["properties"]["sheetId"]
    chart_requests: list[dict] = []
    for idx, spec in enumerate(specs):
        req = _build_chart_request_google(
            spec, sheet_by_name, extents, viz_sheet_id, idx,
        )
        if req is not None:
            chart_requests.append(req)

    if chart_requests:
        time.sleep(throttle)
        sheets_service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": chart_requests},
        ).execute()
        logger.technical(
            f"Created {len(chart_requests)} chart(s) on '{VIZ_SHEET}' sheet."
        )


def _sheet_row_count(sheets_service, spreadsheet_id: str, sheet_name: str, throttle: float) -> int:
    """Return the number of rows with values in column A of the given sheet."""
    import time
    time.sleep(throttle)
    try:
        result = sheets_service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{sheet_name}'!A:A",
        ).execute()
        return len(result.get("values", []))
    except Exception:
        return 0


def _build_chart_request_google(
    spec: dict, sheet_by_name: dict, extents: dict,
    viz_sheet_id: int, chart_index: int,
) -> dict | None:
    """
    Build a single `addChart` batchUpdate request from a chart spec.

    Returns None if the source data sheet has no data rows.

    Google Sheets chart specs use 0-indexed row/col boundaries, and
    require sheetId rather than sheet name. Anchor positions are also
    specified as sheet-relative cell coordinates.
    """
    table_name = spec["table"]
    table_meta = sheet_by_name.get(table_name)
    if not table_meta:
        return None
    table_sheet_id = table_meta["properties"]["sheetId"]
    last_row = extents.get(table_name, 1)
    if last_row < 2:
        return None

    cat_info = spec["categories"]
    cat_first = cat_info["first_row"]
    cat_last = cat_info["last_row"] if cat_info["last_row"] is not None else last_row
    cat_last = min(cat_last, last_row)

    # Convert 1-indexed to 0-indexed for API
    def grid_range(col: int, first_row: int, last_row: int) -> dict:
        return {
            "sheetId": table_sheet_id,
            "startRowIndex": first_row - 1,
            "endRowIndex": last_row,      # end is exclusive
            "startColumnIndex": col - 1,
            "endColumnIndex": col,
        }

    # Parse anchor cell like "A1" or "J20"
    anchor = spec["anchor"]
    import re
    m = re.match(r"([A-Z]+)(\d+)", anchor)
    if not m:
        anchor_col, anchor_row = 0, 0
    else:
        col_letters, row_num = m.group(1), int(m.group(2))
        anchor_col = 0
        for ch in col_letters:
            anchor_col = anchor_col * 26 + (ord(ch) - ord("A") + 1)
        anchor_col -= 1  # 0-indexed
        anchor_row = row_num - 1

    kind = spec["kind"]
    if kind == "pie":
        chart_spec_body = {
            "title": spec["title"],
            "pieChart": {
                "legendPosition": "RIGHT_LEGEND",
                "domain": {
                    "domainSource": {
                        "sources": [grid_range(
                            cat_info["col"], cat_first, cat_last
                        )]
                    }
                },
                "series": {
                    "sourceRange": {
                        "sources": [grid_range(
                            spec["series"][0]["col"], cat_first, cat_last
                        )]
                    }
                },
            },
        }
    else:
        # column_grouped or column_stacked
        stacked_type = "STACKED" if kind == "column_stacked" else "NONE"
        series_list = []
        for s in spec["series"]:
            series_list.append({
                "series": {
                    "sourceRange": {
                        "sources": [grid_range(s["col"], 1, cat_last)]
                    }
                },
                "targetAxis": "LEFT_AXIS",
            })
        chart_spec_body = {
            "title": spec["title"],
            "basicChart": {
                "chartType": "COLUMN",
                "legendPosition": "BOTTOM_LEGEND",
                "stackedType": stacked_type,
                "headerCount": 1,  # First row of each series range is the label
                "axis": [
                    {"position": "BOTTOM_AXIS", "title": spec.get("x_axis", "")},
                    {"position": "LEFT_AXIS", "title": spec.get("y_axis", "")},
                ],
                "domains": [{
                    "domain": {
                        "sourceRange": {
                            "sources": [grid_range(
                                cat_info["col"], cat_first, cat_last
                            )]
                        }
                    }
                }],
                "series": series_list,
            },
        }

    return {
        "addChart": {
            "chart": {
                "spec": chart_spec_body,
                "position": {
                    "overlayPosition": {
                        "anchorCell": {
                            "sheetId": viz_sheet_id,
                            "rowIndex": anchor_row,
                            "columnIndex": anchor_col,
                        },
                        "widthPixels": 600,
                        "heightPixels": 371,
                    }
                },
            }
        }
    }


def _cells_to_values_array(cells: list) -> list[list]:
    """
    Convert a list of (row, col, value) tuples to a 2D values array
    suitable for Google Sheets batch update. Empty cells are filled with "".
    """
    if not cells:
        return []
    max_row = max(c[0] for c in cells)
    max_col = max(c[1] for c in cells)
    values: list[list] = [["" for _ in range(max_col)] for _ in range(max_row)]
    for row, col, value in cells:
        values[row - 1][col - 1] = value
    return values
