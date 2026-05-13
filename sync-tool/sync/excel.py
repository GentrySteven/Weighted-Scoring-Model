"""
Excel Module

Handles all Excel spreadsheet operations using openpyxl.
Uses shared utility functions for column letter conversion.
Respects protected columns with special handling for Month Completed.
"""

import platform
import re
import subprocess
import time
from pathlib import Path
from typing import Any, Optional

from sync.config_manager import ConfigManager
from sync.logging_manager import LoggingManager
from sync.utils import col_letter
from sync.validation import SpreadsheetValidator

SYNC_HEADER_COLOR = "B8CCE4"


class ExcelError(Exception):
    """Raised when an Excel operation fails."""
    pass


class ExcelManager:
    """
    Manages all Excel spreadsheet operations.

    Operations can be performed individually (each one loads and saves the
    workbook) or grouped inside a transaction block. Inside a transaction,
    the workbook is loaded once on entry, kept in memory while multiple
    write operations are applied, and saved exactly once on exit. This is
    much faster for multi-step syncs.

    Usage without a transaction (each call re-opens and saves the file):
        manager.write_rows(headers, rows)
        manager.delete_row(5)
        manager.update_row(headers, row_data, 10)
        # Three full load/save cycles.

    Usage with a transaction (single load, single save):
        with manager.transaction():
            manager.write_rows(headers, rows)
            manager.delete_row(5)
            manager.update_row(headers, row_data, 10)
        # One load at block entry, one save at block exit.

    If an exception is raised inside the transaction block, the workbook
    is NOT saved — the on-disk file is left untouched, preserving the
    pre-transaction state.
    """

    def __init__(self, config: ConfigManager, logger: LoggingManager):
        self.config = config
        self.logger = logger
        self.validator = SpreadsheetValidator(config, logger)
        self.target_dir = Path(config.get("excel", "target_directory", default="") or "")
        self.spreadsheet_name = config.get_spreadsheet_name()
        self.file_path = self.target_dir / f"{self.spreadsheet_name}.xlsx"

        # Transaction state. When non-None, helper methods use this
        # in-memory workbook instead of loading from disk on every call.
        # The active transaction is committed (saved) when the context
        # manager exits cleanly.
        self._active_wb = None
        self._transaction_depth = 0

    # -------------------------------------------------------------------------
    # Transaction support
    # -------------------------------------------------------------------------

    def transaction(self):
        """
        Return a context manager that groups multiple workbook operations
        into a single load-save cycle.

        Nested transaction() calls are allowed and are treated as a single
        outermost transaction — only the outermost `__exit__` saves the
        workbook. If any exception is raised inside the block, the save
        is skipped and the on-disk file is left unchanged.

        Returns:
            A context manager object.
        """
        manager = self

        class _Transaction:
            def __enter__(self):
                if manager._transaction_depth == 0:
                    openpyxl = manager._ensure_openpyxl()
                    manager._active_wb = openpyxl.load_workbook(str(manager.file_path))
                manager._transaction_depth += 1
                return manager

            def __exit__(self, exc_type, exc_val, exc_tb):
                manager._transaction_depth -= 1
                if manager._transaction_depth == 0:
                    wb = manager._active_wb
                    manager._active_wb = None
                    if exc_type is None and wb is not None:
                        # No exception — save the accumulated changes
                        wb.save(str(manager.file_path))
                # Do not suppress exceptions; let them propagate
                return False

        return _Transaction()

    def _load_or_active(self, data_only: bool = False):
        """
        Return the active transaction workbook if one exists, otherwise
        load a fresh workbook from disk.

        Callers that get an ad-hoc (non-transaction) workbook are
        responsible for saving it themselves. Callers that get the
        transaction workbook must NOT save it — the transaction manager
        saves exactly once on exit.

        Returns:
            (workbook, is_from_transaction) — is_from_transaction is True
            when the caller received the active transaction workbook
            and must not save it.
        """
        if self._active_wb is not None:
            return self._active_wb, True
        openpyxl = self._ensure_openpyxl()
        # data_only mode is only honored for fresh loads; transactions
        # always use the full workbook (with formulas) because we may
        # need to write formulas during the same transaction.
        wb = openpyxl.load_workbook(str(self.file_path), data_only=data_only)
        return wb, False

    def _ensure_openpyxl(self):
        """Ensure openpyxl is available."""
        try:
            import openpyxl
            return openpyxl
        except ImportError:
            raise ExcelError(
                "openpyxl is not installed. Run:\n"
                "  pip install archivesspace-accession-sync[excel]"
            )

    def file_exists(self) -> bool:
        return self.file_path.exists()

    def get_file_path(self) -> Path:
        return self.file_path

    def is_file_locked(self) -> tuple[bool, str]:
        """Check if the file is locked by another process."""
        lock_file = self.file_path.parent / f"~${self.file_path.name}"
        if lock_file.exists():
            return True, self._identify_lock_holder()

        try:
            with open(self.file_path, "a"):
                pass
            return False, ""
        except (IOError, PermissionError):
            return True, self._identify_lock_holder()

    def _identify_lock_holder(self) -> str:
        """Attempt to identify which process holds the lock."""
        system = platform.system()
        try:
            if system == "Windows":
                result = subprocess.run(
                    ["powershell", "-Command",
                     f"Get-Process | Where-Object {{$_.MainWindowTitle -like '*{self.spreadsheet_name}*'}}"],
                    capture_output=True, text=True, timeout=5,
                )
                if result.stdout.strip():
                    return f"Locked by: {result.stdout.strip()}"
            elif system in ("Linux", "Darwin"):
                result = subprocess.run(
                    ["lsof", str(self.file_path)],
                    capture_output=True, text=True, timeout=5,
                )
                if result.stdout.strip():
                    lines = result.stdout.strip().split("\n")
                    if len(lines) > 1:
                        parts = lines[1].split()
                        if len(parts) >= 2:
                            return f"Locked by: {parts[0]} (PID: {parts[1]})"
        except (FileNotFoundError, subprocess.TimeoutExpired, Exception):
            pass
        return "Locked by another application"

    def wait_for_unlock(self) -> bool:
        """Wait for the file to become unlocked with retries."""
        max_retries = self.config.get("retry", "file_lock_retries", default=5)
        interval = self.config.get("retry", "file_lock_interval", default=60)

        for attempt in range(max_retries):
            is_locked, lock_info = self.is_file_locked()
            if not is_locked:
                return True

            remaining = max_retries - attempt - 1
            self.logger.summary(f"File locked. {lock_info}. Retrying in {interval}s ({remaining} left).")
            print(f"\n  File is locked. {lock_info}")
            print(f"  Retrying in {interval}s ({remaining} retries remaining).\n")
            time.sleep(interval)

        return False

    def create_spreadsheet(self, headers: list[str]) -> Path:
        """Create a new spreadsheet with headers and formatting."""
        openpyxl = self._ensure_openpyxl()
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Accession Data and Scores"

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            if header.startswith("[Sync]"):
                cell.fill = PatternFill(start_color=SYNC_HEADER_COLOR, end_color=SYNC_HEADER_COLOR, fill_type="solid")

        self._auto_size_columns(ws)

        # Checkbox data validation for format columns
        format_keywords = self.config.get("format_keywords", default={})
        for col_idx, header in enumerate(headers, 1):
            if header in format_keywords:
                dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
                cl = col_letter(col_idx)
                dv.add(f"{cl}2:{cl}1048576")
                ws.add_data_validation(dv)

        # Hidden vocabulary sheets
        self._create_vocabulary_sheets(wb)

        # Scoring criteria sheet (embedded mode only — linked-workbook mode
        # uses a separate file that the user must create)
        scoring_mode = self.config.get_data(
            "scoring_criteria", "excel_scoring_mode", default="embedded_sheet"
        )
        if scoring_mode == "embedded_sheet":
            self._create_scoring_criteria_sheet(wb)

        # Supporting sheets (empty structure, populated during sync)
        self._create_supporting_sheets(wb)

        self.target_dir.mkdir(parents=True, exist_ok=True)
        wb.save(str(self.file_path))
        self.logger.summary(f"Created new Excel spreadsheet: {self.file_path.name}")
        return self.file_path

    def read_data(self) -> tuple[list[str], list[dict]]:
        """Read all data from the spreadsheet."""
        wb, in_transaction = self._load_or_active(data_only=True)
        ws = wb["Accession Data and Scores"]

        headers = [ws.cell(row=1, column=c).value or "" for c in range(1, ws.max_column + 1)]

        rows: list[dict] = []
        for row_idx in range(2, ws.max_row + 1):
            row_data: dict[str, Any] = {}
            has_data = False
            for col_idx, header in enumerate(headers, 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    has_data = True
                row_data[header] = val
            row_data["accession_id"] = row_data.get("Accession ID")
            if has_data:
                rows.append(row_data)

        if not in_transaction:
            wb.close()
        return headers, rows

    def get_column_map(self) -> dict[str, int]:
        """Get the current column name to 1-indexed position mapping."""
        wb, in_transaction = self._load_or_active(data_only=True)
        ws = wb["Accession Data and Scores"]
        column_map: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val:
                column_map[val] = col_idx
        if not in_transaction:
            wb.close()
        return column_map

    def write_rows(
        self, headers: list[str], rows: list[dict],
        start_row: int = 2, is_completion_event: bool = False,
    ) -> None:
        """
        Write rows to the spreadsheet with column-type-aware handling.

        Each column is handled differently based on its type:
        1. Protected columns — skipped if they already have a value
           (scoring formulas, manually assigned fields). Exception:
           Month Completed is writable during completion events.
        2. Formula columns — the formula string is written (not a value)
           using dynamic column references from the current column map.
        3. Subject Descriptor columns — populated from the _subject_descriptors
           list in the row data, matched by column number.
        4. Sync columns — populated from the _sync_data dict, with [Sync] Status
           built from the _changes list or _is_new flag.
        5. Regular data columns — written directly from the row dict.

        Args:
            headers: List of column header strings (defines column order).
            rows: List of row data dicts to write.
            start_row: First row to write (1-indexed, default 2 = after headers).
            is_completion_event: If True, Month Completed is writable.
        """
        wb, in_transaction = self._load_or_active()
        ws = wb["Accession Data and Scores"]

        # Build column map from actual spreadsheet headers so formulas
        # reference the correct positions regardless of column order
        column_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}

        for row_offset, row_data in enumerate(rows):
            row_idx = start_row + row_offset

            for col_idx, header in enumerate(headers, 1):
                # --- Protected columns ---
                # Skip if the column is protected AND already has a value.
                # This preserves scoring formulas, user notes, and manual fields.
                if self.validator.is_protected_column(header, is_completion_event=is_completion_event):
                    existing = ws.cell(row=row_idx, column=col_idx).value
                    if existing is not None:
                        continue

                # --- Formula columns ---
                # Write the formula (e.g., =HYPERLINK(...), =COUNTIF(...))
                # using dynamic column references from the column map
                formula = self.validator.get_column_formula(header, row_idx, column_map)
                if formula:
                    ws.cell(row=row_idx, column=col_idx, value=formula)
                    continue

                # --- Subject Descriptor columns ---
                # These are populated from a list, mapped by the column number
                # (e.g., "Subject Descriptor (#3)" gets the 3rd descriptor)
                if header.startswith("Subject Descriptor"):
                    descriptors = row_data.get("_subject_descriptors", [])
                    match = re.search(r"#(\d+)", header)
                    sd_num = int(match.group(1)) if match else 0
                    if sd_num and sd_num <= len(descriptors):
                        ws.cell(row=row_idx, column=col_idx, value=descriptors[sd_num - 1])
                    continue

                # --- Sync tracking columns ---
                # [Sync] Status is built from change detection results;
                # all other sync columns store lock_versions, IDs, or values
                if header.startswith("[Sync]"):
                    sync_data = row_data.get("_sync_data", {})
                    if header == "[Sync] Status":
                        changes = row_data.get("_changes", [])
                        if changes:
                            status = "Updated — " + ", ".join(changes)
                        elif row_data.get("_is_new"):
                            status = "New"
                        else:
                            status = "Up to date"
                        ws.cell(row=row_idx, column=col_idx, value=status)
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=sync_data.get(header, ""))
                    continue

                # --- Regular data columns ---
                value = row_data.get(header)
                if value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=value)

        self._auto_size_columns(ws)
        if not in_transaction:
            wb.save(str(self.file_path))
        self.logger.technical(f"Wrote {len(rows)} rows to spreadsheet.")

    def update_row(self, headers: list[str], row_data: dict, row_idx: int, **kwargs) -> None:
        """Update a single row."""
        self.write_rows(headers, [row_data], start_row=row_idx, **kwargs)

    def delete_row(self, row_idx: int) -> None:
        """Delete a row from the spreadsheet."""
        wb, in_transaction = self._load_or_active()
        ws = wb["Accession Data and Scores"]
        ws.delete_rows(row_idx)
        if not in_transaction:
            wb.save(str(self.file_path))
        self.logger.technical(f"Deleted row {row_idx}.")

    def clear_data(self) -> None:
        """Clear all data rows, preserving headers."""
        wb, in_transaction = self._load_or_active()
        ws = wb["Accession Data and Scores"]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        if not in_transaction:
            wb.save(str(self.file_path))
        self.logger.summary("Cleared all data from spreadsheet (headers preserved).")

    def find_row_by_accession_id(self, accession_id: int) -> Optional[int]:
        """Find the row number for a given accession ID."""
        wb, in_transaction = self._load_or_active(data_only=True)
        ws = wb["Accession Data and Scores"]

        acc_id_col = None
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).value == "Accession ID":
                acc_id_col = c
                break

        if acc_id_col is None:
            if not in_transaction:
                wb.close()
            return None

        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=acc_id_col).value
            if val is not None and int(val) == accession_id:
                if not in_transaction:
                    wb.close()
                return row_idx

        if not in_transaction:
            wb.close()
        return None

    def _auto_size_columns(self, ws) -> None:
        """Auto-size columns based on content width."""
        for column_cells in ws.columns:
            max_length = 0
            col_letter_val = None
            for cell in column_cells:
                if col_letter_val is None:
                    col_letter_val = cell.column_letter
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            if col_letter_val and max_length > 0:
                ws.column_dimensions[col_letter_val].width = min(max_length + 2, 50)

    def _create_vocabulary_sheets(self, wb) -> None:
        """Create hidden sheets for structured vocabularies."""
        vocab_sheets = [
            "Approved Subject Descriptors",
            "Access Issues Vocabulary",
            "Conservation Issues Vocabulary",
            "Digital Issues Vocabulary",
            "Documentation Issues Options",
            "Other Processing Options",
            "Physical Space Options",
            "Processing Project Types",
        ]
        for name in vocab_sheets:
            ws = wb.create_sheet(title=name)
            ws.sheet_state = "hidden"

        # Populate defaults
        ws = wb["Documentation Issues Options"]
        for idx, opt in enumerate(self.config.get("documentation_use_issues_options", default=[]), 1):
            ws.cell(row=idx, column=1, value=opt)

        ws = wb["Processing Project Types"]
        for idx, pt in enumerate(self.config.get("processing_project_types", default=[]), 1):
            ws.cell(row=idx, column=1, value=pt)

    def _create_supporting_sheets(self, wb) -> None:
        """
        Create supporting sheets that are populated during sync operations.

        Creates four types of supporting sheets:
        - Backlog Change Over Time: monthly trend data
        - Processing Projects Over Time: monthly completion data
        - Backlog At a Glance: status summary dashboard
        - Processing Queue sheets: one per configured queue

        These sheets are created empty during initial workbook creation
        and populated by the sync engine on each run.
        """
        from sync.processing_queue import (
            BacklogAtAGlanceBuilder, queue_sheet_name,
        )

        # Backlog Change Over Time (existing trend tracking)
        ws = wb.create_sheet(title="Backlog Change Over Time")
        headers = ["Month and Year", "Physical Backlog (Linear Feet)", "Digital Backlog (GB)"]
        for idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=idx, value=h)

        # Processing Projects Over Time (existing completion tracking)
        ws = wb.create_sheet(title="Processing Projects Over Time")
        ws.cell(row=1, column=1, value="Month and Year")

        # Backlog At a Glance (snapshot dashboard, populated during sync)
        wb.create_sheet(title=BacklogAtAGlanceBuilder.SHEET_NAME)

        # Processing Queue sheets (one per configured queue)
        # Sheets are created empty here; the sync engine populates them
        # by computing the queue from current spreadsheet rows.
        queue_config = self.config.get_data("processing_queue", default={})
        queues = queue_config.get("queues", [])
        for queue in queues:
            sheet_name = queue_sheet_name(queue.get("name", "Queue"))
            wb.create_sheet(title=sheet_name)
            self.logger.technical(f"Created processing queue sheet: {sheet_name}")

        # Visualization sheets (three hidden data tables + one visible
        # charts sheet). The data tables are populated during sync with
        # computed aggregations; chart objects are added to the charts
        # sheet by the visualization writer.
        from sync.visualizations import (
            VIZ_SHEET, VIZ_MONTHLY_TABLE, VIZ_COMPLETION_TABLE, VIZ_SUBJECTS_TABLE,
        )
        for hidden_name in (VIZ_MONTHLY_TABLE, VIZ_COMPLETION_TABLE, VIZ_SUBJECTS_TABLE):
            sheet = wb.create_sheet(title=hidden_name)
            sheet.sheet_state = "hidden"
        wb.create_sheet(title=VIZ_SHEET)
        self.logger.technical(
            "Created visualization sheets (3 hidden data tables + Visualizations)."
        )

    def _create_scoring_criteria_sheet(self, wb) -> None:
        """
        Create the embedded "Scoring Criteria" sheet from data.yml.

        Reads the configured scoring dimensions from data.yml and uses
        the ScoringCriteriaBuilder to populate a new sheet matching the
        original Scoring Criteria spreadsheet structure (sections for
        quantitative and strategic factors, score tables for each
        dimension, and a weights table at the bottom).

        The sheet name is "Scoring Criteria - DO NOT MOVE" to match the
        original Google Sheets convention.
        """
        from sync.scoring_criteria import write_to_openpyxl_sheet, SHEET_NAME

        criteria = self.config.get_data("scoring_criteria", default={})
        dimensions = criteria.get("dimensions", {})

        if not dimensions:
            self.logger.warning(
                "No scoring criteria configured — skipping scoring sheet creation."
            )
            return

        ws = wb.create_sheet(title=SHEET_NAME)
        write_to_openpyxl_sheet(ws, dimensions)
        self.logger.technical(
            f"Created scoring criteria sheet with {len(dimensions)} dimensions."
        )
